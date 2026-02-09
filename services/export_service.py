"""Export service for generating Excel files."""

import io
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config.logging_config import get_logger
from database.session import get_db
from database.repositories.vacancy_repository import VacancyRepository
from database.repositories.filter_repository import FilterRepository
from utils.helpers import format_date

logger = get_logger(__name__)


class ExportService:
    """Service for exporting vacancies to Excel."""

    def __init__(self) -> None:
        """Initialize export service."""
        pass

    def export_vacancies_to_excel(
        self,
        user_id: int,
        filter_id: Optional[int] = None,
        limit: Optional[int] = None,
    ) -> io.BytesIO:
        """
        Export vacancies to Excel file.

        Args:
            user_id: User ID
            filter_id: Optional filter ID to filter vacancies
            limit: Optional limit on number of vacancies

        Returns:
            BytesIO object with Excel file content
        """
        try:
            db = next(get_db())
            vacancy_repo = VacancyRepository(db)
            filter_repo = FilterRepository(db)

            # Get vacancies
            if filter_id:
                filter_obj = filter_repo.get_by_id(filter_id)
                if not filter_obj:
                    raise ValueError(f"Filter {filter_id} not found")

                # Search vacancies by filter criteria
                vacancies = vacancy_repo.search(
                    profession=filter_obj.profession,
                    city=filter_obj.city,
                    company_name=filter_obj.company_name,
                    limit=limit or 1000,
                )
                filter_name = filter_obj.name
            else:
                # Get all vacancies for user (from database)
                # Note: In current implementation, vacancies are not user-specific
                # This would need to be changed if we want user-specific exports
                vacancies = vacancy_repo.search(limit=limit or 1000)
                filter_name = "Все вакансии"

            if not vacancies:
                raise ValueError("No vacancies found")

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Вакансии"

            # Define styles
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border_side = Side(style="thin", color="000000")
            border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
            center_alignment = Alignment(horizontal="center", vertical="center")
            wrap_alignment = Alignment(wrap_text=True, vertical="top")

            # Headers
            headers = [
                "№",
                "Дата размещения",
                "Источник",
                "Наименование юридического лица",
                "Адрес юридического лица",
                "Наименование специальности",
                "Количество вакантных мест",
                "Зарплата",
                "Контактное лицо",
                "Номер телефона",
                "Ссылка на вакансию",
            ]

            # Write headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border

            # Write data
            for row_num, vacancy in enumerate(vacancies, 2):
                ws.cell(row=row_num, column=1, value=row_num - 1).border = border  # №
                ws.cell(row=row_num, column=2, value=format_date(vacancy.date_posted)).border = border  # Дата
                ws.cell(row=row_num, column=3, value=vacancy.source or "").border = border  # Источник
                ws.cell(row=row_num, column=4, value=vacancy.company_name or "").border = border  # Компания
                ws.cell(row=row_num, column=5, value=vacancy.company_address or "").border = border  # Адрес
                ws.cell(row=row_num, column=6, value=vacancy.position or "").border = border  # Специальность
                ws.cell(row=row_num, column=7, value=vacancy.vacancies_count or "").border = border  # Количество
                ws.cell(row=row_num, column=8, value=vacancy.salary or "").border = border  # Зарплата
                ws.cell(row=row_num, column=9, value=vacancy.contact_person or "").border = border  # Контакт
                ws.cell(row=row_num, column=10, value=vacancy.contact_phone or "").border = border  # Телефон
                ws.cell(row=row_num, column=11, value=vacancy.url or "").border = border  # Ссылка

                # Set alignment for all cells
                for col_num in range(1, 12):
                    cell = ws.cell(row=row_num, column=col_num)
                    if col_num in [1, 2, 3, 7]:  # Numbers and dates - center
                        cell.alignment = center_alignment
                    else:  # Text - wrap
                        cell.alignment = wrap_alignment

            # Auto-adjust column widths
            column_widths = {
                "A": 6,   # №
                "B": 18,  # Дата
                "C": 15,  # Источник
                "D": 30,  # Компания
                "E": 35,  # Адрес
                "F": 25,  # Специальность
                "G": 12,  # Количество
                "H": 15,  # Зарплата
                "I": 20,  # Контакт
                "J": 18,  # Телефон
                "K": 50,  # Ссылка
            }

            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width

            # Set row height for header
            ws.row_dimensions[1].height = 25

            # Freeze first row
            ws.freeze_panes = "A2"

            # Save to BytesIO
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            logger.info(f"Exported {len(vacancies)} vacancies to Excel (filter: {filter_name})")
            return excel_file

        except Exception as e:
            logger.error(f"Error exporting vacancies to Excel: {e}", exc_info=True)
            raise
